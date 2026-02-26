"""
HD Supply™ Velocity Validator
Developed by: Ben F. Benjamaa

A modern, sophisticated application for validating velocity codes against Snowflake data.

Features:
- Modern GUI with HD Supply™ black and yellow branding
- Automated Snowflake SSO authentication
- Real-time 10-step progress tracking
- VLOOKUP functionality for velocity code matching
- Automated DCSKU column generation (DC + USN concatenation)
- Excel export with HD Supply™ formatting
- Summary statistics sheet
- Match/Mismatch validation with color coding

Requirements:
- Python 3.8+
- pandas, openpyxl, snowflake-connector-python
- HD Supply email for Snowflake authentication

Usage:
1. Run the application
2. Select Excel/CSV file with JDA_ITEM and JDA_LOC columns
3. Enter HD Supply email address
4. Click PROCESS DATA
5. Authenticate via browser (SSO)
6. Review generated Excel report with validation results
"""

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import snowflake.connector
import os
from datetime import datetime
import threading
from tkinter import font as tkfont

class ModernButton(tk.Canvas):
    """
    Custom modern button widget with hover effects for HD Supply™ interface.
    
    Creates a canvas-based button with smooth hover animations and HD Supply branding.
    
    Args:
        parent: Parent widget
        text: Button text to display
        command: Function to execute on click
        bg_color: Normal background color
        fg_color: Text color
        hover_color: Background color on hover
        **kwargs: Additional canvas parameters (width, height)
    """
    def __init__(self, parent, text, command, bg_color, fg_color, hover_color, **kwargs):
        width = kwargs.pop('width', 200)
        height = kwargs.pop('height', 50)
        super().__init__(parent, width=width, height=height, bg=parent['bg'], 
                        highlightthickness=0, **kwargs)
        
        self.command = command
        self.bg_color = bg_color
        self.fg_color = fg_color
        self.hover_color = hover_color
        self.text = text
        
        # Create button rectangle with HD Supply styling
        self.rect = self.create_rectangle(0, 0, width, height, 
                                          fill=bg_color, outline="")
        self.text_item = self.create_text(width/2, height/2, 
                                         text=text, fill=fg_color,
                                         font=("Segoe UI", 12, "bold"))
        
        # Bind mouse events for interactivity
        self.bind("<Button-1>", lambda e: self.command())
        self.bind("<Enter>", self.on_enter)
        self.bind("<Leave>", self.on_leave)
        self.configure(cursor="hand2")
        
    def on_enter(self, e):
        """Change button color on mouse hover"""
        self.itemconfig(self.rect, fill=self.hover_color)
        
    def on_leave(self, e):
        """Restore button color when mouse leaves"""
        self.itemconfig(self.rect, fill=self.bg_color)

class VelocityValidatorApp:
    """
    Main application class for HD Supply™ Velocity Validator.
    
    This application provides a sophisticated GUI for validating velocity codes
    by comparing proposed velocities against current velocities from Snowflake.
    
    Key Features:
    - Automated Snowflake authentication using SSO
    - Real-time progress tracking with 10-step process visualization
    - Data merging on JDA_ITEM and JDA_LOC columns
    - DCSKU column generation (concatenation of DC + USN)
    - Excel output with two sheets: detailed data and summary statistics
    - HD Supply™ branded formatting with color-coded matches/mismatches
    """
    def __init__(self, root):
        self.root = root
        self.root.title("HD Supply™ Velocity Validator")
        self.root.geometry("900x750")
        self.root.resizable(False, False)
        
        # Modern HD Supply color scheme - Black background with Yellow accents
        self.bg_black = "#000000"
        self.hd_yellow = "#FFD700"
        self.hd_bright_yellow = "#FFED4E"
        self.dark_gray = "#1A1A1A"
        self.medium_gray = "#2D2D2D"
        self.light_gray = "#404040"
        self.text_gray = "#CCCCCC"
        self.accent_gold = "#FFA500"
        
        # Configure root window
        self.root.configure(bg=self.bg_black)
        
        # Variables
        self.input_file_path = tk.StringVar()
        self.snowflake_data = None
        
        # Configure custom styles
        self.setup_styles()
        self.setup_gui()
        
    def setup_styles(self):
        """Setup custom ttk styles"""
        style = ttk.Style()
        style.theme_use('clam')
        
        # Configure Progressbar
        style.configure("Yellow.Horizontal.TProgressbar",
                       troughcolor=self.medium_gray,
                       background=self.hd_yellow,
                       bordercolor=self.bg_black,
                       lightcolor=self.hd_yellow,
                       darkcolor=self.hd_yellow)
        
    def setup_gui(self):
        # Main container with padding
        main_container = tk.Frame(self.root, bg=self.bg_black)
        main_container.pack(fill="both", expand=True, padx=0, pady=0)
        
        # Header section with company branding
        header_frame = tk.Frame(main_container, bg=self.bg_black, height=120)
        header_frame.pack(fill="x", padx=0, pady=0)
        header_frame.pack_propagate(False)
        
        # Company name with TM
        company_frame = tk.Frame(header_frame, bg=self.bg_black)
        company_frame.pack(pady=(25, 5))
        
        company_label = tk.Label(
            company_frame,
            text="HD SUPPLY",
            font=("Segoe UI", 36, "bold"),
            bg=self.bg_black,
            fg=self.hd_bright_yellow
        )
        company_label.pack(side="left")
        
        tm_label = tk.Label(
            company_frame,
            text="™",
            font=("Segoe UI", 18, "bold"),
            bg=self.bg_black,
            fg=self.hd_bright_yellow
        )
        tm_label.pack(side="left", anchor="n", padx=(2, 0))
        
        # Title with enhanced styling
        title_label = tk.Label(
            header_frame,
            text="VELOCITY VALIDATOR",
            font=("Segoe UI", 20, "bold"),
            bg=self.bg_black,
            fg=self.hd_yellow,
            pady=5
        )
        title_label.pack(pady=(0, 5))
        
        # Separator line
        separator = tk.Frame(main_container, bg=self.hd_yellow, height=2)
        separator.pack(fill="x", padx=20, pady=(0, 20))
        
        # Content frame with padding
        content_frame = tk.Frame(main_container, bg=self.bg_black)
        content_frame.pack(fill="both", expand=True, padx=40, pady=0)
        
        # File selection section
        self.create_section(content_frame, "STEP 1: SELECT FILE", 0)
        
        file_info_frame = tk.Frame(content_frame, bg=self.dark_gray, highlightbackground=self.medium_gray, highlightthickness=1)
        file_info_frame.pack(fill="x", pady=(0, 20))
        
        self.file_label = tk.Label(
            file_info_frame,
            text="📄 No file selected",
            bg=self.dark_gray,
            fg=self.text_gray,
            font=("Segoe UI", 11),
            anchor="w",
            padx=25,
            pady=18
        )
        self.file_label.pack(fill="x")
        
        browse_btn = ModernButton(
            content_frame,
            text="📂 BROWSE FILE",
            command=self.browse_file,
            bg_color=self.hd_yellow,
            fg_color=self.bg_black,
            hover_color=self.hd_bright_yellow,
            width=220,
            height=48
        )
        browse_btn.pack(pady=(12, 25))
        
        # Snowflake connection section
        self.create_section(content_frame, "STEP 2: SNOWFLAKE CONNECTION", 20)
        
        sf_frame = tk.Frame(content_frame, bg=self.dark_gray, highlightbackground=self.medium_gray, highlightthickness=1)
        sf_frame.pack(fill="x", pady=(0, 25), padx=0)
        
        # Email input only
        self.sf_inputs = {}
        
        field_frame = tk.Frame(sf_frame, bg=self.dark_gray)
        field_frame.pack(fill="x", padx=25, pady=20)
        
        label = tk.Label(
            field_frame,
            text="✉️ HD Supply Email:",
            bg=self.dark_gray,
            fg=self.hd_yellow,
            font=("Segoe UI", 11, "bold"),
            width=18,
            anchor="w"
        )
        label.pack(side="left", padx=(0, 10))
        
        entry = tk.Entry(
            field_frame,
            bg=self.medium_gray,
            fg=self.hd_bright_yellow,
            font=("Segoe UI", 11),
            insertbackground=self.hd_yellow,
            relief="solid",
            bd=1,
            highlightbackground=self.light_gray,
            highlightcolor=self.hd_yellow,
            highlightthickness=1
        )
        entry.pack(side="left", fill="x", expand=True, ipady=10, ipadx=12)
        entry.insert(0, "your.email@hdsupply.com")
        self.sf_inputs['email'] = entry
        
        # Info label with icon
        info_label = tk.Label(
            sf_frame,
            text="🔒 Authentication will open in your browser automatically",
            bg=self.dark_gray,
            fg=self.text_gray,
            font=("Segoe UI", 9, "italic"),
            anchor="w"
        )
        info_label.pack(fill="x", padx=25, pady=(0, 18))
        
        # Process button with enhanced styling
        process_btn = ModernButton(
            content_frame,
            text="⚡ PROCESS DATA",
            command=self.process_data,
            bg_color=self.hd_yellow,
            fg_color=self.bg_black,
            hover_color=self.hd_bright_yellow,
            width=280,
            height=60
        )
        process_btn.pack(pady=25)
        self.process_btn = process_btn
        
        # Progress window will be created when processing starts
        self.progress_window = None
        self.progress_bar = None
        self.progress_label = None
        self.step_labels = []
        
        # Footer with gradient effect
        footer_frame = tk.Frame(main_container, bg=self.bg_black, height=50)
        footer_frame.pack(fill="x", side="bottom")
        footer_frame.pack_propagate(False)
        
        footer_separator = tk.Frame(footer_frame, bg=self.hd_yellow, height=2)
        footer_separator.pack(fill="x", padx=20)
        
        # Footer content container
        footer_content = tk.Frame(footer_frame, bg=self.bg_black)
        footer_content.pack(fill="both", expand=True)
        
        # Left side - copyright
        footer_left = tk.Label(
            footer_content,
            text=f"© {datetime.now().year} HD Supply™ | Version 1.0",
            font=("Segoe UI", 8),
            bg=self.bg_black,
            fg=self.text_gray
        )
        footer_left.pack(side="left", padx=30, pady=12)
        
        # Right side - developer credit
        footer_right = tk.Label(
            footer_content,
            text="Developed by: Ben F. Benjamaa",
            font=("Segoe UI", 9, "italic"),
            bg=self.bg_black,
            fg=self.hd_yellow
        )
        footer_right.pack(side="right", padx=30, pady=12)
        
    def create_section(self, parent, title, pady_top):
        """Create a section header with enhanced styling"""
        section_container = tk.Frame(parent, bg=self.bg_black)
        section_container.pack(fill="x", pady=(pady_top, 10))
        
        # Decorative line before title
        left_line = tk.Frame(section_container, bg=self.accent_gold, height=2, width=30)
        left_line.pack(side="left", padx=(0, 10))
        
        section_label = tk.Label(
            section_container,
            text=title,
            font=("Segoe UI", 12, "bold"),
            bg=self.bg_black,
            fg=self.hd_bright_yellow,
            anchor="w"
        )
        section_label.pack(side="left")
        
    def create_progress_window(self):
        """Create a detailed progress tracking window"""
        self.progress_window = tk.Toplevel(self.root)
        self.progress_window.title("Processing...")
        self.progress_window.geometry("600x400")
        self.progress_window.resizable(False, False)
        self.progress_window.configure(bg=self.bg_black)
        self.progress_window.transient(self.root)
        self.progress_window.grab_set()
        
        # Center the window
        self.progress_window.update_idletasks()
        x = (self.progress_window.winfo_screenwidth() // 2) - (600 // 2)
        y = (self.progress_window.winfo_screenheight() // 2) - (400 // 2)
        self.progress_window.geometry(f"600x400+{x}+{y}")
        
        # Header
        header = tk.Label(
            self.progress_window,
            text="⚙️ Processing Velocity Validation",
            font=("Segoe UI", 16, "bold"),
            bg=self.bg_black,
            fg=self.hd_bright_yellow
        )
        header.pack(pady=(20, 10))
        
        # Separator
        separator = tk.Frame(self.progress_window, bg=self.hd_yellow, height=2)
        separator.pack(fill="x", padx=40, pady=(0, 20))
        
        # Steps container
        steps_frame = tk.Frame(self.progress_window, bg=self.bg_black)
        steps_frame.pack(fill="both", expand=True, padx=40, pady=10)
        
        # Define processing steps
        steps = [
            "1. Connecting to Snowflake...",
            "2. Authenticating user...",
            "3. Fetching velocity data...",
            "4. Loading input file...",
            "5. Validating data structure...",
            "6. Merging datasets...",
            "7. Comparing velocities...",
            "8. Generating Excel report...",
            "9. Applying formatting...",
            "10. Saving output file..."
        ]
        
        self.step_labels = []
        for step in steps:
            step_frame = tk.Frame(steps_frame, bg=self.bg_black)
            step_frame.pack(fill="x", pady=3)
            
            icon_label = tk.Label(
                step_frame,
                text="⏳",
                font=("Segoe UI", 10),
                bg=self.bg_black,
                fg=self.text_gray,
                width=3
            )
            icon_label.pack(side="left")
            
            text_label = tk.Label(
                step_frame,
                text=step,
                font=("Segoe UI", 10),
                bg=self.bg_black,
                fg=self.text_gray,
                anchor="w"
            )
            text_label.pack(side="left", fill="x")
            
            self.step_labels.append((icon_label, text_label))
        
        # Progress bar section
        progress_container = tk.Frame(self.progress_window, bg=self.dark_gray)
        progress_container.pack(fill="x", padx=40, pady=20)
        
        self.progress_label = tk.Label(
            progress_container,
            text="Starting...",
            font=("Segoe UI", 9, "italic"),
            bg=self.dark_gray,
            fg=self.hd_yellow
        )
        self.progress_label.pack(pady=(10, 5))
        
        self.progress_bar = ttk.Progressbar(
            progress_container,
            mode="determinate",
            length=500,
            style="Yellow.Horizontal.TProgressbar"
        )
        self.progress_bar.pack(pady=(5, 10))
        self.progress_bar['value'] = 0
        
    def update_progress_step(self, step_index, status="active"):
        """Update a specific step's status
        status: 'active', 'complete', 'error'
        """
        if not self.progress_window or step_index >= len(self.step_labels):
            return
            
        icon_label, text_label = self.step_labels[step_index]
        
        if status == "active":
            icon_label.config(text="⏳", fg=self.hd_yellow)
            text_label.config(fg=self.hd_yellow, font=("Segoe UI", 10, "bold"))
        elif status == "complete":
            icon_label.config(text="✓", fg="#00FF00")
            text_label.config(fg=self.text_gray, font=("Segoe UI", 10))
        elif status == "error":
            icon_label.config(text="✗", fg="#FF0000")
            text_label.config(fg="#FF0000", font=("Segoe UI", 10, "bold"))
            
        # Update progress bar
        if self.progress_bar:
            progress_percent = ((step_index + 1) / len(self.step_labels)) * 100
            self.progress_bar['value'] = progress_percent
            
            if self.progress_label:
                self.progress_label.config(text=f"Progress: {int(progress_percent)}%")
            
        self.progress_window.update()
        
    def close_progress_window(self):
        """Close the progress window"""
        if self.progress_window:
            self.progress_window.grab_release()
            self.progress_window.destroy()
            self.progress_window = None
    
    def browse_file(self):
        """
        Open file dialog for user to select Excel/CSV file for validation.
        
        Displays a file picker with filters for CSV and Excel files.
        Updates the UI to show the selected filename with truncation if needed.
        """
        filename = filedialog.askopenfilename(
            title="Select Excel/CSV file",
            filetypes=[
                ("CSV files", "*.csv"),
                ("Excel files", "*.xlsx *.xls"),
                ("All files", "*.*")
            ]
        )
        if filename:
            self.input_file_path.set(filename)
            display_name = os.path.basename(filename)
            # Truncate long filenames for display
            if len(display_name) > 50:
                display_name = display_name[:47] + "..."
            # Update UI with selected file (yellow text indicates selection)
            self.file_label.config(text=f"✓ {display_name}", fg=self.hd_yellow)
            
    def connect_snowflake(self):
        """
        Establish connection to Snowflake and fetch velocity data.
        
        Uses externalbrowser authentication (SSO) for secure access.
        Queries the SKUEXTRACT table for velocity codes and aliases columns
        to match the expected format in the input file.
        
        Returns:
            bool: True if connection and data fetch successful, False otherwise
        """
        try:
            # Automated connection using externalbrowser authentication (opens browser for SSO)
            con = snowflake.connector.connect(
                user=self.sf_inputs['email'].get().strip(),  # HD Supply email from user input
                account="HDSUPPLY-DATA",  # HD Supply Snowflake account
                authenticator="externalbrowser",  # SSO authentication
                insecure_mode=True  # Allow insecure connections
            )
            
            cur = con.cursor()
            
            # Query Snowflake for velocity data
            # Aliases ITEM -> JDA_ITEM and LOC -> JDA_LOC to match input file format
            query = """
            SELECT
                ITEM as JDA_ITEM,
                LOC as JDA_LOC,
                UDC_VELOCITY_CODE
            FROM
                EDP.STD_JDA.SKUEXTRACT
            """
            
            cur.execute(query)
            
            # Fetch results and convert to DataFrame
            columns = [col[0] for col in cur.description]
            results = cur.fetchall()
            self.snowflake_data = pd.DataFrame(results, columns=columns)
            
            cur.close()
            con.close()
            return True
            
        except Exception as e:
            error_msg = str(e)
            self.root.after(0, lambda: messagebox.showerror(
                "Connection Error", 
                f"Failed to connect to Snowflake:\n\n{error_msg}"
            ))
            return False
            
    def validate_inputs(self):
        """
        Validate user inputs before processing.
        
        Checks that:
        1. An input file has been selected
        2. A valid HD Supply email has been entered
        3. Email contains @hdsupply.com domain
        
        Returns:
            bool: True if all inputs are valid, False otherwise
        """
        # Check if file has been selected
        if not self.input_file_path.get():
            messagebox.showwarning("Missing Input", "Please select an Excel/CSV file!")
            return False
            
        # Validate HD Supply email address
        email = self.sf_inputs['email'].get().strip()
        if not email or email == "your.email@hdsupply.com":
            messagebox.showwarning("Missing Input", "Please enter your HD Supply email address!")
            return False
            
        # Ensure email is from HD Supply domain
        if "@hdsupply.com" not in email.lower():
            messagebox.showwarning("Invalid Input", "Please enter a valid HD Supply email address!")
            return False
                
        return True
            
    def process_data(self):
        """Start data processing in a separate thread"""
        if not self.validate_inputs():
            return
            
        # Run processing in a separate thread to keep UI responsive
        thread = threading.Thread(target=self.process_data_thread, daemon=True)
        thread.start()
        
    def process_data_thread(self):
        """Process the data in a background thread"""
        try:
            # Create progress window
            self.root.after(0, self.create_progress_window)
            import time
            time.sleep(0.3)  # Brief pause to show window
            
            # Step 0: Connecting to Snowflake
            self.root.after(0, lambda: self.update_progress_step(0, "active"))
            time.sleep(0.2)
            
            # Step 1: Authenticating
            self.root.after(0, lambda: self.update_progress_step(0, "complete"))
            self.root.after(0, lambda: self.update_progress_step(1, "active"))
            
            if not self.connect_snowflake():
                self.root.after(0, lambda: self.update_progress_step(1, "error"))
                time.sleep(1)
                self.root.after(0, self.close_progress_window)
                return
            
            # Step 2: Fetching velocity data (completed in connect_snowflake)
            self.root.after(0, lambda: self.update_progress_step(1, "complete"))
            self.root.after(0, lambda: self.update_progress_step(2, "active"))
            time.sleep(0.3)
            self.root.after(0, lambda: self.update_progress_step(2, "complete"))
            
            # Step 3: Loading input file
            self.root.after(0, lambda: self.update_progress_step(3, "active"))
            file_path = self.input_file_path.get()
            if file_path.endswith('.csv'):
                df = pd.read_csv(file_path)
            else:
                df = pd.read_excel(file_path)
            self.root.after(0, lambda: self.update_progress_step(3, "complete"))
            
            # Step 4: Validating data structure
            self.root.after(0, lambda: self.update_progress_step(4, "active"))
            time.sleep(0.2)
            
            # Validate columns
            if 'JDA_ITEM' not in df.columns or 'JDA_LOC' not in df.columns:
                self.root.after(0, lambda: self.update_progress_step(4, "error"))
                time.sleep(0.5)
                self.root.after(0, self.close_progress_window)
                self.root.after(0, lambda: messagebox.showerror(
                    "Column Error",
                    "Required columns JDA_ITEM and/or JDA_LOC not found in input file!"
                ))
                return
            
            self.root.after(0, lambda: self.update_progress_step(4, "complete"))
            
            # Step 5: Merging datasets
            self.root.after(0, lambda: self.update_progress_step(5, "active"))
            
            # Convert merge columns to string type to ensure compatibility
            df['JDA_ITEM'] = df['JDA_ITEM'].astype(str)
            df['JDA_LOC'] = df['JDA_LOC'].astype(str)
            self.snowflake_data['JDA_ITEM'] = self.snowflake_data['JDA_ITEM'].astype(str) # type: ignore
            self.snowflake_data['JDA_LOC'] = self.snowflake_data['JDA_LOC'].astype(str) # type: ignore
            
            df_merged = df.merge(
                self.snowflake_data, # type: ignore
                on=['JDA_ITEM', 'JDA_LOC'],
                how='left'
            )
            
            # Add DCSKU column (concatenate DC + USN)
            if 'DC' in df_merged.columns and 'USN' in df_merged.columns:
                df_merged['DCSKU'] = df_merged['DC'].astype(str) + df_merged['USN'].astype(str)
            
            time.sleep(0.3)
            self.root.after(0, lambda: self.update_progress_step(5, "complete"))
            
            # Step 6: Comparing velocities
            self.root.after(0, lambda: self.update_progress_step(6, "active"))
            
            # Rename UDC_VELOCITY_CODE to Current_Velocity if it exists
            if 'UDC_VELOCITY_CODE' in df_merged.columns:
                df_merged.rename(columns={'UDC_VELOCITY_CODE': 'Current_Velocity'}, inplace=True)
            else:
                # Column not found, create empty Current_Velocity column
                df_merged['Current_Velocity'] = None
            
            if 'PROPOSED_VELOCITY' in df_merged.columns and 'Current_Velocity' in df_merged.columns:
                df_merged['Match'] = df_merged.apply(
                    lambda row: row['Current_Velocity'] == row['PROPOSED_VELOCITY'] 
                    if pd.notna(row['Current_Velocity']) and pd.notna(row['PROPOSED_VELOCITY'])
                    else False,
                    axis=1
                )
            else:
                if 'PROPOSED_VELOCITY' not in df_merged.columns:
                    self.root.after(0, lambda: messagebox.showwarning(
                        "Warning",
                        "PROPOSED_VELOCITY column not found in input file.\nMatch column will be set to False."
                    ))
                df_merged['Match'] = False
            
            time.sleep(0.2)
            self.root.after(0, lambda: self.update_progress_step(6, "complete"))
            
            # Step 7: Generating Excel report
            self.root.after(0, lambda: self.update_progress_step(7, "active"))
            
            # Generate output filename
            input_dir = os.path.dirname(file_path)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"Velocity_Validated_{timestamp}.xlsx"
            output_path = os.path.join(input_dir, output_filename)
            
            time.sleep(0.2)
            self.root.after(0, lambda: self.update_progress_step(7, "complete"))
            
            # Step 8: Applying formatting
            self.root.after(0, lambda: self.update_progress_step(8, "active"))
            time.sleep(0.3)
            
            # Step 9: Saving output file
            self.root.after(0, lambda: self.update_progress_step(8, "complete"))
            self.root.after(0, lambda: self.update_progress_step(9, "active"))
            
            # Save to Excel with formatting
            self.save_formatted_excel(df_merged, output_path)
            
            time.sleep(0.3)
            self.root.after(0, lambda: self.update_progress_step(9, "complete"))
            
            # Calculate statistics
            total_rows = len(df_merged)
            matches = df_merged['Match'].sum() if 'Match' in df_merged.columns else 0
            mismatches = total_rows - matches
            
            # Close progress window
            time.sleep(0.5)
            self.root.after(0, self.close_progress_window)
            
            self.root.after(0, lambda: messagebox.showinfo(
                "Processing Complete",
                f"✓ Velocity validation completed successfully!\n\n"
                f"Output: {output_filename}\n\n"
                f"Statistics:\n"
                f"• Total Records: {total_rows:,}\n"
                f"• Matches: {matches:,}\n"
                f"• Mismatches: {mismatches:,}\n\n"
                f"Columns Added:\n"
                f"• Current_Velocity (from Snowflake)\n"
                f"• Match (True/False comparison)"
            ))
            
        except Exception as e:
            import time
            error_message = str(e)
            time.sleep(0.5)
            self.root.after(0, self.close_progress_window)
            self.root.after(0, lambda msg=error_message: messagebox.showerror(
                "Processing Error",
                f"An error occurred during processing:\n\n{msg}"
            ))
            
    def save_formatted_excel(self, df, output_path):
        """Save DataFrame to Excel with HD Supply formatting and Summary sheet"""
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Write main data sheet
            df.to_excel(writer, sheet_name='Velocity Validation', index=False)
            
            # Create Summary sheet with statistics
            total_records = len(df)
            matches = df['Match'].sum() if 'Match' in df.columns else 0
            mismatches = total_records - matches
            
            summary_data = {
                'Statistics': ['Total Records', 'Matches', 'Mismatches'],
                'Count': [total_records, int(matches), int(mismatches)]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Get the workbook and worksheets
            workbook = writer.book
            worksheet = writer.sheets['Velocity Validation']
            summary_sheet = writer.sheets['Summary']
            
            # Apply formatting
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            
            # HD Supply color scheme
            header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            header_font = Font(color="FFD700", bold=True, size=11)
            
            yellow_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
            red_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            green_fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
            
            thin_border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            
            # Format headers
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
                
            # Highlight Match column
            if 'Match' in df.columns:
                match_col_idx = df.columns.get_loc('Match') + 1
                for row_idx in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row_idx, column=match_col_idx)
                    cell_value = cell.value
                    
                    if cell_value == False or cell_value == 'False':
                        cell.fill = red_fill
                        cell.font = Font(color="CC0000", bold=True)
                    elif cell_value == True or cell_value == 'True':
                        cell.fill = green_fill
                        cell.font = Font(color="006600", bold=True)
                        
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
                    
            # Highlight Current_Velocity column
            if 'Current_Velocity' in df.columns:
                vel_col_idx = df.columns.get_loc('Current_Velocity') + 1
                for row_idx in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row_idx, column=vel_col_idx)
                    cell.fill = yellow_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
                    
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 3, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Format Summary sheet
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            
            # HD Supply branded header
            header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            header_font = Font(color="FFD700", bold=True, size=12)
            
            data_font = Font(size=11, bold=True)
            number_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
            
            thin_border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            
            # Format summary headers
            for cell in summary_sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            
            # Format summary data rows
            for row_idx in range(2, 5):  # Rows 2-4 (Total, Matches, Mismatches)
                for col_idx in range(1, 3):  # Columns A-B
                    cell = summary_sheet.cell(row=row_idx, column=col_idx)
                    cell.font = data_font
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    if col_idx == 2:  # Count column
                        cell.fill = number_fill
                        cell.number_format = '#,##0'
            
            # Set column widths for summary
            summary_sheet.column_dimensions['A'].width = 20
            summary_sheet.column_dimensions['B'].width = 15
            
            # Add title row
            summary_sheet.insert_rows(1)
            summary_sheet['A1'] = 'VELOCITY VALIDATION SUMMARY'
            summary_sheet.merge_cells('A1:B1')
            title_cell = summary_sheet['A1']
            title_cell.font = Font(color="FFD700", bold=True, size=14)
            title_cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
                


def main():
    root = tk.Tk()
    
    # Center window on screen
    window_width = 900
    window_height = 750
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    center_x = int(screen_width/2 - window_width/2)
    center_y = int(screen_height/2 - window_height/2)
    root.geometry(f'{window_width}x{window_height}+{center_x}+{center_y}')
    
    app = VelocityValidatorApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
