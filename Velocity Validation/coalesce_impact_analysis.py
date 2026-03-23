"""
HDS Velocity Reclassification — Impact Analysis: Prior Logic vs. COALESCE Logic
Generates a standalone Excel workbook comparing the old (March 4) and new (March 19) runs.
Focuses on SKUs that transitioned OUT of the New/Forecasted bucket under updated logic.
"""

import pandas as pd
import numpy as np
import os
import sys
from datetime import datetime
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# ── File paths ────────────────────────────────────────────────────────────────
OLD_PATH = r"C:\Users\1015723\OneDrive - HD Supply, Inc\Desktop\New Velocity Reclassification\Updated-HDS_SKU_Velocity_Reclassification_Analysis_MARCH26.xlsx"
NEW_PATH = r"C:\Users\1015723\OneDrive - HD Supply, Inc\Desktop\New Velocity Reclassification\Updated-HDS_SKU_Velocity_Reclassification_Analysis_New Logic.xlsx"
OUT_DIR  = r"C:\Users\1015723\OneDrive - HD Supply, Inc\Desktop\New Velocity Reclassification"
OUT_FILE = os.path.join(OUT_DIR, "HDS_Velocity_COALESCE_Impact_Analysis.xlsx")

SHEET = "Velocity Calculation HDS"
KEY_COLS = ["JDA_ITEM", "JDA_LOC"]

# Columns to carry into the detail tab
DETAIL_COLS = [
    "JDA_ITEM", "JDA_LOC", "DC", "MCAT", "SKU_DESCRIPTION",
    "IS_FORECASTED", "SKU_STATUS", "FIRST_RECEIPT_FLG", "SUPERSEDED_SKU",
    "SYSTEM_VELOCITY", "NEW_PROPOSED_VELOCITY", "SERVICE_LEVEL",
    "VELOCITY_CHANGE_CLASS", "VELOCITY_REASON",
    "TOTAL_FORECASTED_QTY", "WEEKLY_AVERAGE_FORECASTED_QTY",
    "TOTAL_FORECASTED_DOLLARS", "WEEKLY_AVERAGE_FORECASTED_DOLLARS",
    "VELOCITY_WEIGHT", "CALCULATED_VELOCITY_PERCENTILE",
    "CURRENT_SSCOV", "PROPOSED_SSCOV",
    "CURRENT_SAFETY_STOCK", "PROPOSED_SAFETY_STOCK",
    "SAFETY_STOCK_CHANGE", "SAFETY_STOCK_CHANGE_DOLLARS",
    "COGS",
]


def add_excel_table(writer, df, sheet_name, table_name, max_col_width=45):
    """Add an Excel Table (ListObject) with filters, banded rows, and auto-fit columns."""
    ws = writer.sheets[sheet_name]
    if len(df) == 0:
        return
    n_rows = len(df)
    n_cols = len(df.columns)
    end_col = get_column_letter(n_cols)
    ref = f"A1:{end_col}{n_rows + 1}"
    style = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False
    )
    tbl = Table(displayName=table_name, ref=ref)
    tbl.tableStyleInfo = style
    ws.add_table(tbl)
    for i, col in enumerate(df.columns, 1):
        col_letter = get_column_letter(i)
        header_len = len(str(col))
        max_data = df[col].astype(str).str.len().max() if n_rows > 0 else 0
        width = min(max(header_len, max_data) + 3, max_col_width)
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = "A2"


def create_pivot_tables(filepath, detail_rows):
    """Open the workbook in Excel via COM and create real interactive PivotTables."""
    import win32com.client as win32
    import pythoncom

    pythoncom.CoInitialize()
    xl = None
    try:
        xl = win32.DispatchEx("Excel.Application")
        xl.Visible = False
        xl.DisplayAlerts = False
        xl.ScreenUpdating = False

        wb = xl.Workbooks.Open(os.path.abspath(filepath))

        # Excel constants
        xlDatabase = 1
        xlPivotTableVersion15 = 5
        xlRowField = 1
        xlColumnField = 2
        xlDataField = 4
        xlCount = -4112
        xlSum = -4157
        xlTopToBottom = 1

        pivot_counter = [0]

        def add_pivot(source_sheet, dest_name, row_fields, col_fields, data_fields):
            pivot_counter[0] += 1
            src_ws = wb.Sheets(source_sheet)
            used = src_ws.UsedRange
            src_range = src_ws.Range(
                src_ws.Cells(1, 1),
                src_ws.Cells(used.Rows.Count, used.Columns.Count)
            )
            dest_ws = wb.Sheets.Add(After=wb.Sheets(wb.Sheets.Count))
            dest_ws.Name = dest_name
            cache = wb.PivotCaches().Create(
                SourceType=xlDatabase,
                SourceData=src_range,
                Version=xlPivotTableVersion15
            )
            pt_name = f"PT_{pivot_counter[0]}"
            pt = cache.CreatePivotTable(
                TableDestination=dest_ws.Cells(3, 1),
                TableName=pt_name,
                DefaultVersion=xlPivotTableVersion15
            )
            pt.RowAxisLayout(xlTopToBottom)
            for f in row_fields:
                pf = pt.PivotFields(f)
                pf.Orientation = xlRowField
            for f in col_fields:
                pf = pt.PivotFields(f)
                pf.Orientation = xlColumnField
            for f_name, f_func, f_caption in data_fields:
                pf = pt.PivotFields(f_name)
                pf.Orientation = xlDataField
                pf.Function = f_func
                try:
                    pf.Caption = f_caption
                except Exception:
                    pass
            pt.TableStyle2 = "PivotStyleMedium9"
            print(f"  Created PivotTable: {dest_name}")

        if detail_rows > 0:
            src = "Transitioned Detail"

            add_pivot(src, "Pivot - Velocity Shift",
                      ["OLD_PROPOSED_VELOCITY"], ["NEW_PROPOSED_VELOCITY"],
                      [("JDA_ITEM", xlCount, "Count of SKUs")])

            add_pivot(src, "Pivot - By DC",
                      ["DC_NEW"], ["NEW_PROPOSED_VELOCITY"],
                      [("JDA_ITEM", xlCount, "Count of SKUs")])

            add_pivot(src, "Pivot - By MCAT",
                      ["MCAT"], ["NEW_PROPOSED_VELOCITY"],
                      [("JDA_ITEM", xlCount, "Count of SKUs")])

            add_pivot(src, "Pivot - SS Impact by DC",
                      ["DC_NEW"], [],
                      [("OLD_PROPOSED_SS", xlSum, "Sum Old SS"),
                       ("NEW_PROPOSED_SS", xlSum, "Sum New SS"),
                       ("SS_CHANGE_DELTA", xlSum, "Sum SS Delta")])

            add_pivot(src, "Pivot - SS Impact by MCAT",
                      ["MCAT"], [],
                      [("OLD_PROPOSED_SS", xlSum, "Sum Old SS"),
                       ("NEW_PROPOSED_SS", xlSum, "Sum New SS"),
                       ("SS_CHANGE_DELTA", xlSum, "Sum SS Delta")])

            add_pivot(src, "Pivot - Reason x Velocity",
                      ["NEW_VELOCITY_REASON"], ["NEW_PROPOSED_VELOCITY"],
                      [("JDA_ITEM", xlCount, "Count of SKUs")])

        wb.Save()
        wb.Close(False)
        print("  PivotTables saved successfully.")

    except Exception as e:
        print(f"  ERROR creating PivotTables: {e}", file=sys.stderr)
        if xl:
            try:
                xl.Workbooks.Close()
            except Exception:
                pass
    finally:
        if xl:
            xl.Quit()
        pythoncom.CoUninitialize()


def main():
    print("Reading old analysis...")
    old = pd.read_excel(OLD_PATH, sheet_name=SHEET)
    print(f"  Old rows: {len(old):,}")

    print("Reading new analysis...")
    new = pd.read_excel(NEW_PATH, sheet_name=SHEET)
    print(f"  New rows: {len(new):,}")

    # ── Merge on JDA_ITEM + JDA_LOC ──────────────────────────────────────────
    merged = old.merge(new, on=KEY_COLS, how="outer", suffixes=("_OLD", "_NEW"), indicator=True)
    print(f"  Merged rows: {len(merged):,}")
    print(f"  both: {(merged['_merge']=='both').sum():,}  |  old_only: {(merged['_merge']=='left_only').sum():,}  |  new_only: {(merged['_merge']=='right_only').sum():,}")

    # ── FOCUS: Records that were New+Forecasted in OLD ────────────────────────
    # These are the ones potentially impacted by the COALESCE change
    was_new_fcst = merged[
        (merged["_merge"].isin(["both", "left_only"]))
        & (merged["SKU_STATUS_OLD"] == "New")
        & (merged["IS_FORECASTED_OLD"] == "Forecasted")
    ].copy()
    print(f"\n  Records that were New+Forecasted in OLD: {len(was_new_fcst):,}")

    # Classify transition
    def classify_transition(row):
        if row["_merge"] == "left_only":
            return "Dropped from universe"
        new_status = row["SKU_STATUS_NEW"]
        new_fcst   = row["IS_FORECASTED_NEW"]
        if new_status == "New" and new_fcst == "Forecasted":
            return "Stayed New+Forecasted"
        elif new_status == "Not New" and new_fcst == "Forecasted":
            return "Transitioned to Not New+Forecasted"
        elif new_status == "New" and new_fcst == "Not Forecasted":
            return "Changed to New+Not Forecasted"
        elif new_status == "Not New" and new_fcst == "Not Forecasted":
            return "Transitioned to Not New+Not Forecasted"
        else:
            return "Other"

    was_new_fcst["TRANSITION"] = was_new_fcst.apply(classify_transition, axis=1)

    # ── TAB 1: Executive Summary ─────────────────────────────────────────────
    print("\nBuilding executive summary...")

    summary_rows = []

    # Overall universe
    summary_rows.append(("Universe Overview", "", ""))
    summary_rows.append(("Total DC-SKUs (Old — March 4)", f"{len(old):,}", ""))
    summary_rows.append(("Total DC-SKUs (New — March 19)", f"{len(new):,}", ""))
    summary_rows.append(("Net change in universe", f"{len(new) - len(old):+,}", ""))
    summary_rows.append(("", "", ""))

    # New+Forecasted bucket shift
    old_new_fcst = len(old[(old["SKU_STATUS"] == "New") & (old["IS_FORECASTED"] == "Forecasted")])
    new_new_fcst = len(new[(new["SKU_STATUS"] == "New") & (new["IS_FORECASTED"] == "Forecasted")])
    summary_rows.append(("New+Forecasted Bucket Shift", "", ""))
    summary_rows.append(("New+Forecasted (Old)", f"{old_new_fcst:,}", ""))
    summary_rows.append(("New+Forecasted (New)", f"{new_new_fcst:,}", ""))
    summary_rows.append(("Net change", f"{new_new_fcst - old_new_fcst:+,}", f"{(new_new_fcst - old_new_fcst) / old_new_fcst * 100:+.1f}%"))
    summary_rows.append(("", "", ""))

    # Transition breakdown
    summary_rows.append(("Transition Breakdown (of Old New+Forecasted records)", "", ""))
    trans_counts = was_new_fcst["TRANSITION"].value_counts()
    for t in ["Stayed New+Forecasted", "Transitioned to Not New+Forecasted",
              "Changed to New+Not Forecasted", "Transitioned to Not New+Not Forecasted",
              "Dropped from universe", "Other"]:
        cnt = trans_counts.get(t, 0)
        pct = cnt / len(was_new_fcst) * 100 if len(was_new_fcst) > 0 else 0
        summary_rows.append((f"  {t}", f"{cnt:,}", f"{pct:.1f}%"))
    summary_rows.append(("", "", ""))

    # Velocity shift for transitioned records
    transitioned = was_new_fcst[was_new_fcst["TRANSITION"] == "Transitioned to Not New+Forecasted"]
    summary_rows.append(("Velocity Assignments for Transitioned Records (Not New+Forecasted)", "", ""))
    summary_rows.append(("(These were previously forced to C as New+Forecasted)", "", ""))
    if len(transitioned) > 0:
        vel_new = transitioned["NEW_PROPOSED_VELOCITY_NEW"].value_counts().sort_index()
        for v in ["A", "B", "C", "D", "E"]:
            cnt = vel_new.get(v, 0)
            pct = cnt / len(transitioned) * 100 if len(transitioned) > 0 else 0
            summary_rows.append((f"  Now classified as {v}", f"{cnt:,}", f"{pct:.1f}%"))
        summary_rows.append(("  Total transitioned", f"{len(transitioned):,}", ""))
    summary_rows.append(("", "", ""))

    # Safety stock impact for transitioned
    if len(transitioned) > 0:
        ss_old = transitioned["PROPOSED_SAFETY_STOCK_OLD"].sum()
        ss_new = transitioned["PROPOSED_SAFETY_STOCK_NEW"].sum()
        ssd_old = transitioned["SAFETY_STOCK_CHANGE_DOLLARS_OLD"].sum()
        ssd_new = transitioned["SAFETY_STOCK_CHANGE_DOLLARS_NEW"].sum()
        summary_rows.append(("Safety Stock Impact (Transitioned Records Only)", "", ""))
        summary_rows.append((f"  Old proposed safety stock (units)", f"{ss_old:,.0f}", ""))
        summary_rows.append((f"  New proposed safety stock (units)", f"{ss_new:,.0f}", ""))
        summary_rows.append((f"  Delta (units)", f"{ss_new - ss_old:+,.0f}", ""))
        summary_rows.append((f"  Old SS change ($)", f"${ssd_old:,.2f}", ""))
        summary_rows.append((f"  New SS change ($)", f"${ssd_new:,.2f}", ""))
        summary_rows.append(("", "", ""))

    # Full velocity distribution comparison
    summary_rows.append(("Full Velocity Distribution Comparison", "", ""))
    summary_rows.append(("Velocity", "Old Count", "New Count"))
    for v in ["A", "B", "C", "D", "E"]:
        o = len(old[old["NEW_PROPOSED_VELOCITY"] == v])
        n = len(new[new["NEW_PROPOSED_VELOCITY"] == v])
        summary_rows.append((v, f"{o:,}", f"{n:,}"))
    summary_rows.append(("Total", f"{len(old):,}", f"{len(new):,}"))
    summary_rows.append(("", "", ""))

    # Reason distribution comparison
    summary_rows.append(("Velocity Reason Distribution Comparison", "", ""))
    summary_rows.append(("Reason", "Old Count", "New Count"))
    all_reasons = set(old["VELOCITY_REASON"].unique()) | set(new["VELOCITY_REASON"].unique())
    for r in sorted(all_reasons):
        o = len(old[old["VELOCITY_REASON"] == r])
        n = len(new[new["VELOCITY_REASON"] == r])
        summary_rows.append((r, f"{o:,}", f"{n:,}"))

    df_summary = pd.DataFrame(summary_rows, columns=["Metric", "Value", "Detail"])

    # ── TAB 2: Transition Detail (formerly New+Forecasted → Not New+Forecasted)
    print("Building transition detail...")

    # Build clean detail columns from merged
    detail_rows = []
    for _, row in transitioned.iterrows():
        detail_rows.append({
            "JDA_ITEM": row["JDA_ITEM"],
            "JDA_LOC": row["JDA_LOC"],
            "DC_OLD": row.get("DC_OLD"),
            "DC_NEW": row.get("DC_NEW"),
            "MCAT": row.get("MCAT_OLD", row.get("MCAT_NEW")),
            "SKU_DESCRIPTION": row.get("SKU_DESCRIPTION_OLD", row.get("SKU_DESCRIPTION_NEW")),
            "OLD_SKU_STATUS": row.get("SKU_STATUS_OLD"),
            "NEW_SKU_STATUS": row.get("SKU_STATUS_NEW"),
            "OLD_IS_FORECASTED": row.get("IS_FORECASTED_OLD"),
            "NEW_IS_FORECASTED": row.get("IS_FORECASTED_NEW"),
            "OLD_FIRST_RECEIPT_FLG": row.get("FIRST_RECEIPT_FLG_OLD"),
            "NEW_FIRST_RECEIPT_FLG": row.get("FIRST_RECEIPT_FLG_NEW"),
            "SUPERSEDED_SKU": row.get("SUPERSEDED_SKU_OLD", row.get("SUPERSEDED_SKU_NEW")),
            "SYSTEM_VELOCITY": row.get("SYSTEM_VELOCITY_OLD", row.get("SYSTEM_VELOCITY_NEW")),
            "OLD_PROPOSED_VELOCITY": row.get("NEW_PROPOSED_VELOCITY_OLD"),
            "NEW_PROPOSED_VELOCITY": row.get("NEW_PROPOSED_VELOCITY_NEW"),
            "VELOCITY_SHIFT": f"{row.get('NEW_PROPOSED_VELOCITY_OLD', '?')} -> {row.get('NEW_PROPOSED_VELOCITY_NEW', '?')}",
            "OLD_VELOCITY_REASON": row.get("VELOCITY_REASON_OLD"),
            "NEW_VELOCITY_REASON": row.get("VELOCITY_REASON_NEW"),
            "OLD_VELOCITY_CHANGE_CLASS": row.get("VELOCITY_CHANGE_CLASS_OLD"),
            "NEW_VELOCITY_CHANGE_CLASS": row.get("VELOCITY_CHANGE_CLASS_NEW"),
            "NEW_VELOCITY_WEIGHT": row.get("VELOCITY_WEIGHT_NEW"),
            "NEW_VELOCITY_PERCENTILE": row.get("CALCULATED_VELOCITY_PERCENTILE_NEW"),
            "TOTAL_FORECASTED_QTY_NEW": row.get("TOTAL_FORECASTED_QTY_NEW"),
            "WEEKLY_AVG_FCST_QTY_NEW": row.get("WEEKLY_AVERAGE_FORECASTED_QTY_NEW"),
            "WEEKLY_AVG_FCST_DOLLARS_NEW": row.get("WEEKLY_AVERAGE_FORECASTED_DOLLARS_NEW"),
            "OLD_PROPOSED_SSCOV": row.get("PROPOSED_SSCOV_OLD"),
            "NEW_PROPOSED_SSCOV": row.get("PROPOSED_SSCOV_NEW"),
            "OLD_PROPOSED_SS": row.get("PROPOSED_SAFETY_STOCK_OLD"),
            "NEW_PROPOSED_SS": row.get("PROPOSED_SAFETY_STOCK_NEW"),
            "SS_CHANGE_DELTA": (row.get("PROPOSED_SAFETY_STOCK_NEW") or 0) - (row.get("PROPOSED_SAFETY_STOCK_OLD") or 0),
            "OLD_SS_CHANGE_DOLLARS": row.get("SAFETY_STOCK_CHANGE_DOLLARS_OLD"),
            "NEW_SS_CHANGE_DOLLARS": row.get("SAFETY_STOCK_CHANGE_DOLLARS_NEW"),
            "COGS": row.get("COGS_NEW", row.get("COGS_OLD")),
        })

    df_detail = pd.DataFrame(detail_rows)

    # ── TAB 3: Velocity Shift Matrix (Old Proposed → New Proposed for transitioned records)
    print("Building velocity shift matrix...")

    vel_matrix = transitioned.groupby(
        ["NEW_PROPOSED_VELOCITY_OLD", "NEW_PROPOSED_VELOCITY_NEW"]
    ).size().reset_index(name="COUNT")
    vel_matrix.columns = ["OLD_PROPOSED_VELOCITY", "NEW_PROPOSED_VELOCITY", "COUNT"]

    # Pivot
    vel_pivot = vel_matrix.pivot_table(
        index="OLD_PROPOSED_VELOCITY",
        columns="NEW_PROPOSED_VELOCITY",
        values="COUNT",
        fill_value=0,
        aggfunc="sum"
    )
    # Ensure all velocity tiers exist
    for v in ["A", "B", "C", "D", "E"]:
        if v not in vel_pivot.columns:
            vel_pivot[v] = 0
        if v not in vel_pivot.index:
            vel_pivot.loc[v] = 0
    vel_pivot = vel_pivot.reindex(index=["A", "B", "C", "D", "E"], columns=["A", "B", "C", "D", "E"], fill_value=0)
    vel_pivot["Total"] = vel_pivot.sum(axis=1)
    vel_pivot.loc["Total"] = vel_pivot.sum(axis=0)
    vel_pivot.index.name = "Old Proposed \\ New Proposed"
    df_vel_matrix = vel_pivot.reset_index()

    # ── TAB 4: Summary by DC ─────────────────────────────────────────────────
    print("Building DC summary...")

    dc_summary = transitioned.groupby("DC_NEW").agg(
        records_transitioned=("JDA_ITEM", "size"),
        now_A=("NEW_PROPOSED_VELOCITY_NEW", lambda x: (x == "A").sum()),
        now_B=("NEW_PROPOSED_VELOCITY_NEW", lambda x: (x == "B").sum()),
        now_C=("NEW_PROPOSED_VELOCITY_NEW", lambda x: (x == "C").sum()),
        now_D=("NEW_PROPOSED_VELOCITY_NEW", lambda x: (x == "D").sum()),
        now_E=("NEW_PROPOSED_VELOCITY_NEW", lambda x: (x == "E").sum()),
        ss_delta_units=("PROPOSED_SAFETY_STOCK_NEW", lambda x: x.sum() - transitioned.loc[x.index, "PROPOSED_SAFETY_STOCK_OLD"].sum()),
    ).reset_index()
    dc_summary.columns = ["DC", "Records Transitioned", "Now A", "Now B", "Now C", "Now D", "Now E", "SS Delta (Units)"]
    dc_summary = dc_summary.sort_values("Records Transitioned", ascending=False)

    # ── TAB 5: Summary by MCAT ───────────────────────────────────────────────
    print("Building MCAT summary...")

    mcat_col = "MCAT_NEW" if "MCAT_NEW" in transitioned.columns else "MCAT_OLD"
    mcat_summary = transitioned.groupby(mcat_col).agg(
        records_transitioned=("JDA_ITEM", "size"),
        now_A=("NEW_PROPOSED_VELOCITY_NEW", lambda x: (x == "A").sum()),
        now_B=("NEW_PROPOSED_VELOCITY_NEW", lambda x: (x == "B").sum()),
        now_C=("NEW_PROPOSED_VELOCITY_NEW", lambda x: (x == "C").sum()),
        now_D=("NEW_PROPOSED_VELOCITY_NEW", lambda x: (x == "D").sum()),
        now_E=("NEW_PROPOSED_VELOCITY_NEW", lambda x: (x == "E").sum()),
    ).reset_index()
    mcat_summary.columns = ["MCAT", "Records Transitioned", "Now A", "Now B", "Now C", "Now D", "Now E"]
    mcat_summary = mcat_summary.sort_values("Records Transitioned", ascending=False)

    # ── TAB 6: All transitions (full breakdown, not just Not New+Forecasted)
    print("Building full transition breakdown...")

    all_trans = was_new_fcst.groupby("TRANSITION").agg(
        count=("JDA_ITEM", "size"),
    ).reset_index()
    all_trans["% of Total"] = (all_trans["count"] / all_trans["count"].sum() * 100).round(1)
    all_trans = all_trans.sort_values("count", ascending=False)
    all_trans.columns = ["Transition Type", "Record Count", "% of Total"]

    # ── WRITE TO EXCEL ────────────────────────────────────────────────────────
    print(f"\nWriting to {OUT_FILE}...")

    with pd.ExcelWriter(OUT_FILE, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="Executive Summary", index=False)
        df_detail.to_excel(writer, sheet_name="Transitioned Detail", index=False)
        df_vel_matrix.to_excel(writer, sheet_name="Velocity Shift Matrix", index=False)
        dc_summary.to_excel(writer, sheet_name="Summary by DC", index=False)
        mcat_summary.to_excel(writer, sheet_name="Summary by MCAT", index=False)
        all_trans.to_excel(writer, sheet_name="Transition Breakdown", index=False)

        add_excel_table(writer, df_summary, "Executive Summary", "ExecSummary")
        add_excel_table(writer, df_detail, "Transitioned Detail", "TransitionedDetail")
        add_excel_table(writer, df_vel_matrix, "Velocity Shift Matrix", "VelocityShiftMatrix")
        add_excel_table(writer, dc_summary, "Summary by DC", "SummaryByDC")
        add_excel_table(writer, mcat_summary, "Summary by MCAT", "SummaryByMCAT")
        add_excel_table(writer, all_trans, "Transition Breakdown", "TransitionBreakdown")

    # Create real interactive PivotTables via Excel COM
    print("Creating PivotTables via Excel COM automation...")
    create_pivot_tables(OUT_FILE, len(df_detail))

    print(f"Done! Output: {OUT_FILE}")
    print(f"\nKey metrics:")
    print(f"  Old New+Forecasted: {old_new_fcst:,}")
    print(f"  New New+Forecasted: {new_new_fcst:,}")
    print(f"  Transitioned to Not New+Forecasted: {len(transitioned):,}")
    if len(transitioned) > 0:
        vel_dist = transitioned["NEW_PROPOSED_VELOCITY_NEW"].value_counts().sort_index()
        for v, c in vel_dist.items():
            print(f"    Now {v}: {c:,}")


if __name__ == "__main__":
    main()
