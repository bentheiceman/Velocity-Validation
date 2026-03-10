"""Generate a segmented 'not updated' velocity workbook matching the channel-management template.

Template format
- Sheets: A, B, C, D, E, Summary
- Each velocity sheet has header row 1 with columns:
  JDA_ITEM, JDA_LOC, PROPOSED_VELOCITY, PROPOSED_VELOCITY_, SERVICE_LEVEL,
  SAP_VELOCITY, VELOCITY_REASON, Old_Proposed_Velocity

This script compares:
- Thursday recommendations (CSV/XLSX) containing JDA_ITEM/JDA_LOC/PROPOSED_VELOCITY/etc
vs
- Parm Management Weekly Report (XLSX) sheet 'TW Data' containing ITEM/DC NUMBER/VELOCITY

and outputs only rows that are not yet updated in the system:
- Missing in Parm (left_only)
- Present in Parm but VELOCITY != PROPOSED_VELOCITY

Usage
  python parm_velocity_delta_segmented.py \
    --thursday <recs.csv> \
    --parm <parm.xlsx> \
    --out <out.xlsx> \
    --template <template.xlsx>
"""

from __future__ import annotations

import argparse
import shutil
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


DEFAULT_PARM_SHEET = "TW Data"
DEFAULT_TEMPLATE = Path(r"C:\Users\1015723\Downloads\HDS_velocity_changes_segmented_New SKU.xlsx")


def _read_thursday(path: Path) -> pd.DataFrame:
    if path.suffix.lower() in {".xlsx", ".xls"}:
        df = pd.read_excel(path, dtype=str)
    elif path.suffix.lower() == ".csv":
        df = pd.read_csv(path, dtype=str)
    else:
        raise ValueError(f"Unsupported Thursday file type: {path.suffix}")

    df.columns = [str(c).strip() for c in df.columns]
    required = ["JDA_ITEM", "JDA_LOC"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise KeyError(f"Thursday file missing required columns: {missing}")

    df["JDA_ITEM"] = df["JDA_ITEM"].astype(str).str.strip()
    df["JDA_LOC"] = df["JDA_LOC"].astype(str).str.strip()

    # Normalize core columns.
    if "PROPOSED_VELOCITY" not in df.columns:
        candidates = [c for c in df.columns if "PROPOSED" in c.upper() and "VELOC" in c.upper()]
        if not candidates:
            raise KeyError("Thursday file missing PROPOSED_VELOCITY (or a similar proposed velocity column).")
        df = df.rename(columns={candidates[0]: "PROPOSED_VELOCITY"})

    if "PROPOSED_VELOCITY_" not in df.columns:
        # Keep a duplicate column if it doesn't exist.
        df["PROPOSED_VELOCITY_"] = df["PROPOSED_VELOCITY"]

    # Ensure these exist (blank if not provided).
    for col in ["SERVICE_LEVEL", "SAP_VELOCITY", "VELOCITY_REASON"]:
        if col not in df.columns:
            df[col] = ""

    for col in ["PROPOSED_VELOCITY", "PROPOSED_VELOCITY_", "SERVICE_LEVEL", "SAP_VELOCITY", "VELOCITY_REASON"]:
        df[col] = df[col].astype(str).str.strip().replace({"nan": ""})

    return df


def _read_parm(path: Path, sheet: str) -> pd.DataFrame:
    if path.suffix.lower() not in {".xlsx", ".xls"}:
        raise ValueError("Parm report must be an .xlsx/.xls file")

    df = pd.read_excel(path, sheet_name=sheet, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]

    required = ["ITEM", "DC NUMBER", "VELOCITY"]
    col_lookup = {str(c).strip().upper(): str(c).strip() for c in df.columns}
    rename_map: dict[str, str] = {}
    for req in required:
        actual = col_lookup.get(req)
        if actual and actual != req:
            rename_map[actual] = req

    if rename_map:
        df = df.rename(columns=rename_map)

    missing = [c for c in required if c not in df.columns]
    if missing:
        raise KeyError(
            f"Parm sheet '{sheet}' missing required columns: {missing}. "
            "Use the 'TW Data' sheet or pass --sheet."
        )

    df["ITEM"] = df["ITEM"].astype(str).str.strip()
    df["DC NUMBER"] = df["DC NUMBER"].astype(str).str.strip()
    df["VELOCITY"] = df["VELOCITY"].astype(str).str.strip()

    df = df.drop_duplicates(subset=["ITEM", "DC NUMBER"], keep="last")
    return df


def _build_not_updated(thursday: pd.DataFrame, parm: pd.DataFrame) -> pd.DataFrame:
    th = thursday.copy()
    th = th.rename(columns={"JDA_ITEM": "ITEM", "JDA_LOC": "DC NUMBER"})

    parm_keep = parm[["ITEM", "DC NUMBER", "VELOCITY"]].copy()

    merged = th.merge(parm_keep, on=["ITEM", "DC NUMBER"], how="left", indicator=True)

    merged["parm_velocity"] = merged["VELOCITY"].astype(str).str.strip().replace({"nan": ""})
    merged["proposed_velocity"] = merged["PROPOSED_VELOCITY"].astype(str).str.strip().replace({"nan": ""})

    # Not updated if missing in Parm OR mismatch.
    not_updated = merged[(merged["_merge"] == "left_only") | (merged["parm_velocity"] != merged["proposed_velocity"])].copy()

    # Map into template column names.
    out = pd.DataFrame(
        {
            "JDA_ITEM": not_updated["ITEM"].astype(str).str.strip(),
            "JDA_LOC": not_updated["DC NUMBER"].astype(str).str.strip(),
            "PROPOSED_VELOCITY": not_updated["PROPOSED_VELOCITY"].astype(str).str.strip(),
            "PROPOSED_VELOCITY_": not_updated["PROPOSED_VELOCITY_"].astype(str).str.strip(),
            "SERVICE_LEVEL": not_updated["SERVICE_LEVEL"].astype(str).str.strip(),
            # Put today's system velocity in SAP_VELOCITY for validation.
            "SAP_VELOCITY": not_updated["parm_velocity"].astype(str).str.strip(),
            "VELOCITY_REASON": not_updated["VELOCITY_REASON"].astype(str).str.strip(),
            # Preserve what SAP_VELOCITY was at time of Thursday export.
            "Old_Proposed_Velocity": not_updated["SAP_VELOCITY"].astype(str).str.strip(),
        }
    )

    out = out.sort_values(["JDA_LOC", "JDA_ITEM"], kind="stable")
    return out


def _from_existing_delta(delta_xlsx: Path, sheet: str = "not_updated") -> pd.DataFrame:
    """Load the previously generated delta workbook and map into template columns."""
    df = pd.read_excel(delta_xlsx, sheet_name=sheet, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]

    required = ["ITEM", "DC NUMBER", "proposed_velocity", "parm_velocity", "SAP_VELOCITY", "SERVICE_LEVEL", "VELOCITY_REASON"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise KeyError(f"Delta sheet '{sheet}' is missing required columns: {missing}")

    out = pd.DataFrame(
        {
            "JDA_ITEM": df["ITEM"].astype(str).str.strip(),
            "JDA_LOC": df["DC NUMBER"].astype(str).str.strip(),
            "PROPOSED_VELOCITY": df["proposed_velocity"].astype(str).str.strip(),
            "PROPOSED_VELOCITY_": df["proposed_velocity"].astype(str).str.strip(),
            "SERVICE_LEVEL": df["SERVICE_LEVEL"].astype(str).str.strip(),
            "SAP_VELOCITY": df["parm_velocity"].astype(str).str.strip(),
            "VELOCITY_REASON": df["VELOCITY_REASON"].astype(str).str.strip(),
            "Old_Proposed_Velocity": df["SAP_VELOCITY"].astype(str).str.strip(),
        }
    )

    for c in out.columns:
        out[c] = out[c].replace({"nan": ""})

    out = out.sort_values(["JDA_LOC", "JDA_ITEM"], kind="stable")
    return out


def _clear_data_rows(ws) -> None:
    # Preserve header row 1.
    if ws.max_row and ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)


def write_segmented(template_path: Path, out_path: Path, not_updated: pd.DataFrame) -> dict[str, int]:
    out_path.parent.mkdir(parents=True, exist_ok=True)

    # Copy template first to keep formatting, filters, column widths, etc.
    shutil.copyfile(template_path, out_path)

    wb = load_workbook(out_path)

    counts: dict[str, int] = {}

    for v in ["A", "B", "C", "D", "E"]:
        if v not in wb.sheetnames:
            raise KeyError(f"Template is missing expected sheet '{v}'")
        ws = wb[v]
        _clear_data_rows(ws)

        dfv = not_updated[not_updated["PROPOSED_VELOCITY"].astype(str).str.upper().str.strip() == v].copy()
        counts[v] = int(len(dfv))

        if dfv.empty:
            continue

        for row in dfv.itertuples(index=False):
            ws.append(list(row))

    # Update summary counts.
    if "Summary" in wb.sheetnames:
        ws = wb["Summary"]
        # Expect rows 2-6 correspond to A-E.
        for i, v in enumerate(["A", "B", "C", "D", "E"], start=2):
            ws.cell(row=i, column=1).value = v
            ws.cell(row=i, column=2).value = counts.get(v, 0)

    wb.save(out_path)

    return counts


def main() -> int:
    parser = argparse.ArgumentParser(description="Create a segmented not-updated workbook in the standard A-E template format.")
    parser.add_argument("--thursday", help="Thursday recommendations export (.csv/.xlsx)")
    parser.add_argument("--parm", help="Parm Management Weekly Report (.xlsx)")
    parser.add_argument(
        "--delta",
        help="Optional: previously generated parm delta XLSX (uses its 'not_updated' sheet). If provided, --thursday/--parm are not required.",
    )
    parser.add_argument("--out", required=True, help="Output .xlsx path")
    parser.add_argument("--sheet", default=DEFAULT_PARM_SHEET, help=f"Parm sheet name (default: {DEFAULT_PARM_SHEET})")
    parser.add_argument("--delta-sheet", default="not_updated", help="Delta workbook sheet name (default: not_updated)")
    parser.add_argument("--template", default=str(DEFAULT_TEMPLATE), help="Template workbook path")

    args = parser.parse_args()

    out_path = Path(args.out)
    template_path = Path(args.template)

    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    if args.delta:
        delta_path = Path(args.delta)
        if not delta_path.exists():
            raise FileNotFoundError(f"Delta workbook not found: {delta_path}")
        not_updated = _from_existing_delta(delta_path, sheet=args.delta_sheet)
    else:
        if not args.thursday or not args.parm:
            raise ValueError("Either provide --delta, or provide both --thursday and --parm.")

        thursday_path = Path(args.thursday)
        parm_path = Path(args.parm)
        th = _read_thursday(thursday_path)
        parm = _read_parm(parm_path, sheet=args.sheet)
        not_updated = _build_not_updated(th, parm)

    counts = write_segmented(template_path, out_path, not_updated)

    total = int(len(not_updated))
    print("Segmented not-updated workbook created")
    print(f"- not_updated_total_rows: {total}")
    for v in ["A", "B", "C", "D", "E"]:
        print(f"- {v}: {counts.get(v, 0)}")
    print(f"Wrote: {out_path}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
